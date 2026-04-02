from __future__ import annotations
import io
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from .compare import ProductAnalysis
    from parsers.base import ParseError

# ---------------------------------------------------------------------------
# Colour palette (dark-mode-ish, readable on white)
_GREEN  = "FF16A34A"
_YELLOW = "FFCA8A04"
_RED    = "FFDC2626"
_GREY   = "FF6B7280"
_HEADER = "FF1E1E2E"
_ACCENT = "FF6366F1"

_THIN = Side(style="thin", color="FFD1D5DB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header_style(ws, row: int, cols: int) -> None:
    fill = PatternFill("solid", fgColor=_HEADER)
    font = Font(bold=True, color="FFE2E8F0", name="Consolas", size=10)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER


def _row_fill(recommendation: str) -> PatternFill | None:
    mapping = {
        "strong_buy": PatternFill("solid", fgColor="FF052E16"),
        "consider":   PatternFill("solid", fgColor="FF1C1500"),
        "skip":       PatternFill("solid", fgColor="FF1C0000"),
        "not_found":  None,
    }
    return mapping.get(recommendation)


def _rec_label(rec: str) -> str:
    return {
        "strong_buy": "✅ Strong Buy",
        "consider":   "🟡 Consider",
        "skip":       "❌ Skip",
        "not_found":  "⚪ Not Found",
    }.get(rec, rec)


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------

def generate_xlsx(
    analyses: list["ProductAnalysis"],
    errors: list["ParseError"] | None = None,
) -> bytes:
    """Generate a multi-sheet XLSX report and return raw bytes."""
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------
    # Sheet 1 — Opportunities (all products with Skroutz data, best first)
    ws1 = wb.active
    ws1.title = "Opportunities"
    ws1.freeze_panes = "A2"

    hdrs1 = ["Product", "Supplier", "Category", "Barcode",
             "Wholesale €", "Retail €", "Skroutz Low €", "Skroutz High €",
             "Margin €", "Margin %", "Undercut vs RRP %", "Shops", "Rating", "Reviews",
             "Competition", "Score", "Recommendation", "Skroutz URL"]
    ws1.append(hdrs1)
    _header_style(ws1, 1, len(hdrs1))

    for a in analyses:
        p = a.product
        s = a.skroutz
        row = [
            p.name,
            p.source.title(),
            p.category,
            p.barcode,
            p.wholesale_price,
            p.retail_price,
            s.lowest_price if s.found else "",
            s.highest_price if s.found else "",
            a.margin_absolute if s.found else "",
            a.margin_pct if s.found else "",
            a.undercut_vs_retail if s.found else "",
            s.shop_count if s.found else "",
            s.rating if s.found else "",
            s.review_count if s.found else "",
            a.competition_level,
            a.opportunity_score,
            _rec_label(a.recommendation),
            s.product_url,
        ]
        ws1.append(row)
        row_idx = ws1.max_row
        fill = _row_fill(a.recommendation)
        font = Font(name="Consolas", size=9)
        for col in range(1, len(hdrs1) + 1):
            cell = ws1.cell(row=row_idx, column=col)
            if fill:
                cell.fill = fill
            cell.font = font
            cell.border = _BORDER
        # Hyperlink on URL column
        url_cell = ws1.cell(row=row_idx, column=len(hdrs1))
        if s.product_url:
            url_cell.hyperlink = s.product_url
            url_cell.font = Font(name="Consolas", size=9, color=_ACCENT, underline="single")

    _set_col_widths(ws1, [40, 12, 20, 16, 12, 12, 14, 14, 10, 10, 14, 8, 8, 9, 12, 8, 16, 40])

    # ------------------------------------------------------------------
    # Sheet 2 — Not Found (first-mover opportunities)
    ws2 = wb.create_sheet("Not Found")
    ws2.freeze_panes = "A2"
    hdrs2 = ["Product", "Supplier", "Category", "Barcode", "Wholesale €", "Retail €",
             "Markup %", "Note"]
    ws2.append(hdrs2)
    _header_style(ws2, 1, len(hdrs2))

    not_found = [a for a in analyses if not a.skroutz.found]
    for a in not_found:
        p = a.product
        markup = round(((p.retail_price - p.wholesale_price) / p.wholesale_price) * 100, 1) if p.wholesale_price > 0 else 0
        ws2.append([
            p.name, p.source.title(), p.category, p.barcode,
            p.wholesale_price, p.retail_price, markup,
            "No results on Skroutz — potential first-mover opportunity",
        ])
        font = Font(name="Consolas", size=9)
        for col in range(1, len(hdrs2) + 1):
            ws2.cell(ws2.max_row, col).font = font
            ws2.cell(ws2.max_row, col).border = _BORDER

    _set_col_widths(ws2, [40, 12, 20, 16, 12, 12, 10, 50])

    # ------------------------------------------------------------------
    # Sheet 3 — Parse Errors
    ws3 = wb.create_sheet("Parse Errors")
    hdrs3 = ["File", "Row", "Reason", "Raw Data"]
    ws3.append(hdrs3)
    _header_style(ws3, 1, len(hdrs3))

    for e in (errors or []):
        ws3.append([e.filename, e.row, e.reason, e.raw or ""])
        for col in range(1, 5):
            ws3.cell(ws3.max_row, col).border = _BORDER
            ws3.cell(ws3.max_row, col).font = Font(name="Consolas", size=9)

    _set_col_widths(ws3, [30, 6, 60, 80])

    # ------------------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
