from __future__ import annotations
import logging
from pathlib import Path

import openpyxl

from .base import ProductRecord, ParseError

logger = logging.getLogger(__name__)


def _parse_price(value) -> float | None:
    """Convert a cell value to float price, handling both numeric and string forms."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", ".").replace("€", "").strip()
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def parse_xlsx(path: Path) -> tuple[list[ProductRecord], list[ParseError]]:
    """Parse a Bio Tonics-style XLSX pricelist.

    Expected sheet layout (sheet name: 'Table 1'):
      A: ΚΩΔΙΚΟΣ (code)   B: ΠΕΡΙΓΡΑΦΗ (name)
      C: ΧΤ (wholesale)   D: ΠΛΤ (retail)   E: BARCODE

    Category header rows have text only in column A; C/D/E are empty.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    # Try 'Table 1' first, then fall back to first sheet
    if "Table 1" in wb.sheetnames:
        ws = wb["Table 1"]
    else:
        ws = wb.active

    # Derive supplier slug from filename
    name_lower = path.stem.lower()
    if "biotonics" in name_lower or "atcare" in name_lower:
        source = "biotonics"
    else:
        source = path.stem[:20].lower().replace(" ", "_")

    products: list[ProductRecord] = []
    errors: list[ParseError] = []
    current_category = ""

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        code_val, name_val, wh_val, rt_val, bc_val = (row + (None,) * 5)[:5]

        # Category header row: name-like in col A, prices empty
        if code_val is not None and wh_val is None and rt_val is None:
            current_category = str(code_val).strip()
            continue

        # Skip header row that may have slipped through
        if isinstance(code_val, str) and code_val.upper() in ("ΚΩΔΙΚΟΣ", "CODE", "SKU"):
            continue

        if name_val is None:
            continue

        wholesale = _parse_price(wh_val)
        retail = _parse_price(rt_val)

        if wholesale is None or retail is None:
            errors.append(ParseError(
                filename=path.name,
                row=row_idx,
                reason=f"Could not parse prices: XT={wh_val!r} PLT={rt_val!r}",
                raw=str(row),
            ))
            continue

        # Guard against zero prices — would cause division-by-zero in analysis
        if wholesale <= 0 or retail <= 0:
            errors.append(ParseError(
                filename=path.name,
                row=row_idx,
                reason=f"Zero or negative price: wholesale={wholesale} retail={retail}",
                raw=str(row),
            ))
            logger.warning("Row %d in %s has zero/negative price — skipped", row_idx, path.name)
            continue

        # Always store barcode as string; handles int, float, and string cells
        if bc_val is None:
            barcode = ""
        elif isinstance(bc_val, (int, float)):
            barcode = str(int(bc_val))
        else:
            barcode = str(bc_val).strip()

        products.append(ProductRecord(
            source=source,
            code=str(code_val or "").strip(),
            name=str(name_val).strip(),
            wholesale_price=wholesale,
            retail_price=retail,
            barcode=barcode,
            category=current_category,
        ))

    return products, errors
